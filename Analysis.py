import re
import csv
import openpyxl

#Dictionary declared to map state values to total insured values
total_insured_values = {}

#Dictionary declared to map state values to the number of risks
total_number_of_risks = {}

#Dictionary declared to map counties to total insured values
total_insured_values_by_county = {}

#Dictionary declared to determine the 10 postal codes in
#portfolio 1 with the highest average annual loss
most_at_risk_postal_codes_port1 = {}
postal_code_risk_count_p1 = {}

#Dictionary declared to determine the 10 postal codes in
#portfolio 2 with the highest average annual loss
most_at_risk_postal_codes_port2 = {}
postal_code_risk_count_p2 = {}

#Dictionary declared for breakdown of total insured value, risk count,
#and average annual loss by building height band
grouped_data = {}

#Global variables to determine the total AAL of each portfolio
aver_AAL_p1 = 0.0
aver_AAL_p2 = 0.0


#Opens the Datafile
with open('Data_Assessment.csv', 'r') as csvfile:
    counter1 = 0
    counter2 = 0

    reader = csv.reader(csvfile)
    next(reader)
    total_building_value_p1 = 0
    total_building_value_p2 = 0
    total_aal_p1 = 0
    total_aal_p2 = 0
    #Loops through each row in the data file
    for row in reader:

        key_state = row[2]
        key_county = row[3]
        key_postal_code = row[4]
        building_value = int(row[5])
        other_value = int(row[6])
        contents_value = int(row[7])
        time_element_value = int(row[8])

        if row[1] == '1':
            total_building_value_p1 += building_value
            total_aal_p1 += float(row[-1])
        else:
            total_building_value_p2 += building_value
            total_aal_p2 += float(row[-1])

        #This represents the total insured value
        sum_value = building_value + other_value + contents_value + time_element_value

        #This function will determine the total insured value by state.
        #This is done by mapping through the dictionary and adding to 
        #the total insured value, if the state exists in the dictionary.
        def find_insurance_by_state():
            if key_state in total_insured_values:
                total_insured_values[key_state] += sum_value

            else:
                total_insured_values[key_state] = sum_value

        #This function will determine the number of risks by state.
        #This is done by mapping through the dictionary and adding 
        #one for each time a state re-appears.
        def find_risk_by_state():
            if key_state in total_number_of_risks:
                total_number_of_risks[key_state] += 1

            else:
                total_number_of_risks[key_state] = 1
        
        #This function will find the total_insured_values by county with
        #the building code WD10. Each time a county reappears in the data,
        #the total insured value is added onto the sum total by that point
        def find_insurance_by_county():
            if key_county in total_insured_values_by_county:
                if row[10] == 'WD10':
                    total_insured_values_by_county[key_county] += sum_value
            
            else:
                if row[10] == 'WD10':
                    total_insured_values_by_county[key_county] = sum_value
        
        #This function will find the top 10 most at risk postal codes
        #in both portfolio 1 and 2. This is done by determining the
        #total average annual loss for the zipcode.
        def find_postal_codes_at_risk():
            if row[1] == '1':
                if key_postal_code in most_at_risk_postal_codes_port1 and key_postal_code in postal_code_risk_count_p1:
                    most_at_risk_postal_codes_port1[key_postal_code] += float(row[-1])
                    postal_code_risk_count_p1[key_postal_code] += 1

                else:
                    most_at_risk_postal_codes_port1[key_postal_code] = float(row[-1])
                    postal_code_risk_count_p1[key_postal_code] = 1
            elif row[1] == '2':
                if key_postal_code in most_at_risk_postal_codes_port2 and key_postal_code in postal_code_risk_count_p2:
                    most_at_risk_postal_codes_port2[key_postal_code] += float(row[-1])
                    postal_code_risk_count_p2[key_postal_code] += 1

                else:
                    most_at_risk_postal_codes_port2[key_postal_code] = float(row[-1])
                    postal_code_risk_count_p2[key_postal_code] = 1

        #Since the number of stories contains non-digit characters, this
        #function cleans any non-digit characters from the code and then
        #matches the remaining digits with a building range. 
        def categorize_building():
            digits_only = re.sub(r"[^0-9]", "", row[11])

            if (digits_only != ''):
                value = int(digits_only)
            else:
                value = 0

            if value >= 1 and value <= 3:
                return 'Small'
            elif value >= 4 and value <= 7:
                return 'Medium'
            elif value >= 8:
                return 'Large'
            else:
                return 'Unknown'
        

        #This function will group total insured value, risk count, and 
        #average annual loss by building band height if they are in 
        # Pennsylvania.
        def by_band():
            if row[2] == "PA":
                category = categorize_building()
                postal_code = row[4]

                if postal_code not in grouped_data:
                    grouped_data[postal_code] = {
                        'Small': {'Total TIV': 0, 'Risk Count': 0, 'AAL': 0},
                        'Medium': {'Total TIV': 0, 'Risk Count': 0, 'AAL': 0},
                        'Large': {'Total TIV': 0, 'Risk Count': 0, 'AAL': 0},
                        'Unknown': {'Total TIV': 0, 'Risk Count': 0, 'AAL': 0}
                    }
                grouped_data[postal_code][category]['Total TIV'] += sum_value
                grouped_data[postal_code][category]['Risk Count'] += 1
                grouped_data[postal_code][category]['AAL'] += float(row[-1])

        #This function will determine average total AAL for each portfolio
        #and return them on each iteration of the loop to update.
        def average_AAL(aver_AAL_p1, aver_AAL_p2, counter1, counter2):
            if row[1] == '1':
                aver_AAL_p1 += float(row[-1])
                counter1 += 1
            elif row[1] == '2':
                aver_AAL_p2 += float(row[-1])
                counter2 += 1
            return aver_AAL_p1, aver_AAL_p2, counter1, counter2



        find_insurance_by_state()
        find_risk_by_state()
        find_insurance_by_county()
        find_postal_codes_at_risk()
        by_band()
        aver_AAL_p1, aver_AAL_p2, counter1, counter2 = average_AAL(aver_AAL_p1,
                                                       aver_AAL_p2, counter1, 
                                                       counter2)

    print("Averages:")
    print(total_building_value_p1/counter1)
    print(total_building_value_p2/counter2)
    print(total_aal_p1/counter1)
    print(total_aal_p2/counter2)
    print('\n')

#This function will sort the dictionaries to get the desired outputs of data,
#and call a function to write them to an XML file.
def sort_dictionaries():
    largest_five = dict(sorted(total_insured_values_by_county.items(), 
                        key=lambda item: item[1], reverse = True)[:5])

    largest_ten_port1 = dict(sorted(most_at_risk_postal_codes_port1.items(), 
                            key=lambda item: item[1], reverse = True)[:10])
    
    largest_ten_port2 = dict(sorted(most_at_risk_postal_codes_port2.items(), 
                                    key=lambda item: item[1], reverse = True)[:10])

    largest_ten_per_asset_port1 = sort_dictionary_largest_ten(most_at_risk_postal_codes_port1, postal_code_risk_count_p1)
    largest_ten_per_asset_port2 = sort_dictionary_largest_ten(most_at_risk_postal_codes_port2, postal_code_risk_count_p2)



    write_xml(largest_five, largest_ten_port1, largest_ten_port2, largest_ten_per_asset_port1, largest_ten_per_asset_port2)

#This function will find the average annual loss per asset in each zip code
#and return a sorted dictionary with the largest ten in the portfolio
def sort_dictionary_largest_ten(most_at_risk_postal_codes, postal_code_risk_count):
    new_dict = {}
    for key in most_at_risk_postal_codes:
        if key in postal_code_risk_count_p1:
            value1 = most_at_risk_postal_codes[key]
            value2 = postal_code_risk_count[key]
            if value2 != 0: 
                new_dict[key] = value1 / value2
    return dict(sorted(new_dict.items(), key=lambda item: item[1], reverse = True)[:10])




#This writes all the desired data to their respective sheets in an XML workbook.
def write_xml(largest_five, largest_ten_port1, largest_ten_port2, largest_ten_per_asset_port1, largest_ten_per_asset_port2):
    wb = openpyxl.Workbook()

    #Question 1
    ws = wb.active
    ws.title = 'Total Insured Values'
    ws.append(['State', 'Total Insured Value', 'Total Number of Risks'])
    for state in total_insured_values:
        ws.append([state, total_insured_values[state], total_number_of_risks[state]])

    #Question 2
    ws = wb.create_sheet(title='Largest Five')
    ws.append(['County', 'Total Insured Value'])
    for county in largest_five:
        ws.append([county, largest_five[county]])

    #Question 3 Part 1
    ws = wb.create_sheet(title='Largest Ten')
    ws.append(['Portfolio', 'Postal Code', 'Total AAL'])
    for postal_code in largest_ten_port1:
        ws.append([1, postal_code, largest_ten_port1[postal_code]])

    for postal_code in largest_ten_port2:
        ws.append([2, postal_code, largest_ten_port2[postal_code]])

    #Question 3 Part 2
    ws = wb.create_sheet(title='Largest Ten Average')
    ws.append(['Portfolio', 'Postal Code', 'AAL per Project'])
    for postal_code in largest_ten_per_asset_port1:
        ws.append([1, postal_code, largest_ten_per_asset_port1[postal_code]])

    for postal_code in largest_ten_per_asset_port2:
        ws.append([2, postal_code, largest_ten_per_asset_port2[postal_code]])


    #Question 4
    ws = wb.create_sheet(title='Grouped Data')
    ws.append(['Postal Code', 'Small TIV', 'Small Risk Count', 'Small AAL',
    'Medium TIV', 'Medium Risk Count', 'Medium AAL',
    'Large TIV', 'Large Risk Count', 'Large AAL',
    'Unknown TIV', 'Unknown Risk Count', 'Unknown AAL'])

    for postal_code, data in grouped_data.items():
        ws.append([postal_code,
        data['Small']['Total TIV'], data['Small']['Risk Count'], data['Small']['AAL'],
        data['Medium']['Total TIV'], data['Medium']['Risk Count'], data['Medium']['AAL'],
        data['Large']['Total TIV'], data['Large']['Risk Count'], data['Large']['AAL'],
        data['Unknown']['Total TIV'], data['Unknown']['Risk Count'], data['Unknown']['AAL']])

    wb.save('final_workbook3.xlsx')

sort_dictionaries()

print(aver_AAL_p1)
print(aver_AAL_p2)
aver_AAL_p1 = aver_AAL_p1 / counter1 
aver_AAL_p2 = aver_AAL_p2 / counter2
print(counter1)
print(counter2)
print(aver_AAL_p1)
print(aver_AAL_p2)